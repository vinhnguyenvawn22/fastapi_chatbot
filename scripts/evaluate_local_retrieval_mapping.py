from pathlib import Path
import argparse
import asyncio
import json
import os
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT_DIR = Path(__file__).resolve().parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


def _clean(value) -> str:
    return " ".join(str(value or "").split())


def _load_questions(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["groundtruth"]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "") for value in next(rows)]
        question_index = headers.index("question_original")
        id_index = headers.index("gt_id")
        topic_index = headers.index("topic")
        return [
            {
                "gt_id": _clean(row[id_index]),
                "topic": _clean(row[topic_index]),
                "question": _clean(row[question_index]),
            }
            for row in rows
            if row[question_index]
        ]
    finally:
        workbook.close()


def _query_coverage(question: str, docs: list[dict], get_keywords) -> float:
    keywords = set(get_keywords(question))
    if not keywords:
        return 0.0
    evidence = " ".join(
        _clean(
            " ".join(
                str(doc.get(field) or "")
                for field in (
                    "title",
                    "heading",
                    "section_path",
                    "doc_name",
                    "content",
                )
            )
        )
        for doc in docs[:3]
    )
    evidence_terms = set(get_keywords(evidence))
    return round(len(keywords & evidence_terms) / len(keywords), 4)


def _float(doc: dict, field: str):
    value = doc.get(field)
    try:
        return round(float(value), 6) if value is not None else None
    except (TypeError, ValueError):
        return None


def _judge(question: str, docs: list[dict], get_keywords) -> tuple[str, str, float]:
    if not docs:
        return "KHONG_HOP_LY", "Khong tim thay evidence.", 0.0

    top = docs[0]
    coverage = _query_coverage(question, docs, get_keywords)
    vector_score = _float(top, "vector_score") or 0.0
    keyword_score = _float(top, "keyword_score") or 0.0
    rerank_score = _float(top, "rerank_score")
    branches = set(top.get("retrieval_branches") or [])
    signals = 0
    reasons = []

    if coverage >= 0.5:
        signals += 1
        reasons.append(f"do_phu_tu_khoa={coverage:.2f}")
    if vector_score >= 0.45:
        signals += 1
        reasons.append(f"vector={vector_score:.2f}")
    if keyword_score >= 4:
        signals += 1
        reasons.append(f"bm25={keyword_score:.2f}")
    if rerank_score is not None and rerank_score >= 0:
        signals += 1
        reasons.append(f"rerank={rerank_score:.2f}")
    if {"bm25_original", "ann_original"}.issubset(branches):
        signals += 1
        reasons.append("BM25_va_ANN_dong_thuan")

    if signals >= 3:
        return "HOP_LY", ", ".join(reasons), coverage
    if signals >= 1:
        return "CAN_XEM", ", ".join(reasons), coverage
    return "KHONG_HOP_LY", "Khong co tin hieu relevance du nguong.", coverage


async def _evaluate(questions: list[dict]) -> list[dict]:
    from app.data.ambiguity_analyzer import DIRECT_RETRIEVAL
    from app.data.elasticsearch_client import get_keywords, search_documents

    output = []
    for index, item in enumerate(questions, start=1):
        debug = {}
        docs = await search_documents(
            item["question"],
            debug=debug,
            source_type_filter="local_file",
            corpus_filter="local_documents",
            rag_enabled_filter=True,
            exclude_document_names={"PCNTT_MAPPING_FILE.docx"},
            exclude_source_types={"website_uneti", "business_faq_mapping"},
            ambiguity_decision={
                "action": DIRECT_RETRIEVAL,
                "topic": None,
                "confidence": 1.0,
                "reason": "offline_retrieval_evaluation",
                "clarifying_question": None,
            },
        )
        verdict, reason, coverage = _judge(item["question"], docs, get_keywords)
        top = docs[0] if docs else {}
        sources = [
            {
                "doc_name": doc.get("doc_name"),
                "title": doc.get("title"),
                "department": doc.get("department"),
                "document_type": doc.get("document_type"),
                "relative_path": doc.get("relative_path"),
                "rerank_score": _float(doc, "rerank_score"),
                "vector_score": _float(doc, "vector_score"),
                "bm25_score": _float(doc, "bm25_score"),
                "rrf_score": _float(doc, "rrf_score"),
                "preview": _clean(doc.get("content"))[:500],
            }
            for doc in docs[:3]
        ]
        output.append({
            **item,
            "verdict": verdict,
            "reason": reason,
            "query_coverage": coverage,
            "retrieved_count": len(docs),
            "top_doc_name": top.get("doc_name"),
            "top_title": top.get("title"),
            "top_department": top.get("department"),
            "top_document_type": top.get("document_type"),
            "top_rerank_score": _float(top, "rerank_score"),
            "top_vector_score": _float(top, "vector_score"),
            "top_bm25_score": _float(top, "bm25_score"),
            "top_rrf_score": _float(top, "rrf_score"),
            "top_preview": _clean(top.get("content"))[:1000],
            "sources_json": json.dumps(sources, ensure_ascii=False),
            "retrieval_error": "; ".join(
                str(error) for error in debug.get("vector_errors", [])
            ),
        })
        print(f"[{index:02d}/{len(questions)}] {item['gt_id']}: {verdict}")
    return output


def _write_report(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "retrieval_evaluation"
    headers = list(rows[0]) if rows else []
    sheet.append(headers)
    fills = {
        "HOP_LY": PatternFill("solid", fgColor="C6EFCE"),
        "CAN_XEM": PatternFill("solid", fgColor="FFEB9C"),
        "KHONG_HOP_LY": PatternFill("solid", fgColor="FFC7CE"),
    }
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        sheet.append([row.get(header) for header in headers])
        verdict_cell = sheet.cell(sheet.max_row, headers.index("verdict") + 1)
        verdict_cell.fill = fills.get(row["verdict"], PatternFill())

    widths = {
        "A": 16,
        "B": 28,
        "C": 55,
        "D": 18,
        "E": 45,
        "F": 18,
        "G": 14,
        "H": 16,
        "I": 45,
        "J": 50,
        "O": 90,
        "P": 90,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary = workbook.create_sheet("summary")
    counts = {
        verdict: sum(1 for row in rows if row["verdict"] == verdict)
        for verdict in ("HOP_LY", "CAN_XEM", "KHONG_HOP_LY")
    }
    summary.append(["metric", "value"])
    summary.append(["total_questions", len(rows)])
    for verdict, count in counts.items():
        summary.append([verdict, count])
    summary.append(["gemini_calls", 0])
    summary.append(["mapping_source_used_for_scoring", "No"])
    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 18
    workbook.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--questions",
        type=Path,
        default=ROOT_DIR / "groundtruth_pcntt_mapping_80_questions.xlsx",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=ROOT_DIR / "storage" / "local_documents_eval" / "retrieval_80_questions.xlsx",
    )
    args = parser.parse_args()
    questions = _load_questions(args.questions)
    rows = asyncio.run(_evaluate(questions))
    _write_report(args.output, rows)
    counts = {
        verdict: sum(1 for row in rows if row["verdict"] == verdict)
        for verdict in ("HOP_LY", "CAN_XEM", "KHONG_HOP_LY")
    }
    print(json.dumps({"output": str(args.output), **counts}, ensure_ascii=False))


if __name__ == "__main__":
    main()
