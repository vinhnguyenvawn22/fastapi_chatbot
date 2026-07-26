from pathlib import Path
import argparse
import json
import sys

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT_DIR = Path(__file__).resolve().parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

SUPPORT_DOCUMENTS = {
    "2026.03.03.ChatbotAI_CBGV_SV_V4.docx",
    "2026.03.25.AI_HDSD TREN WEB SUPPORT CBGV.docx",
    "2026.03.25.AI_HDSD TREN WEB SUPPORT SV.docx",
}
MANUAL_OVERRIDES = {
    "PCNTT-GT-018": (
        "HOP_LY",
        "Top 3 co huong dan bao hong thiet bi va cach lien he xu ly.",
    ),
    "PCNTT-GT-019": (
        "HOP_LY",
        "Top chunk dung muc Thuc hien bao hong thiet bi phan cung.",
    ),
    "PCNTT-GT-030": (
        "HOP_LY",
        "Top evidence chua cac buoc tham gia khao sat noi bo.",
    ),
    "PCNTT-GT-033": (
        "HOP_LY",
        "Top chunk dung buoc tra loi va nop khao sat ben ngoai.",
    ),
    "PCNTT-GT-054": (
        "HOP_LY",
        "Top chunk dung buoc chon thiet bi trong quy trinh bao hong.",
    ),
    "PCNTT-GT-056": (
        "HOP_LY",
        "Top chunk dung buoc chon linh vuc va ten su co.",
    ),
}


def _rows(path: Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        values = list(sheet.iter_rows(values_only=True))
        headers = list(values[0])
        return [dict(zip(headers, row)) for row in values[1:]]
    finally:
        workbook.close()


def _cosine(left, right) -> float:
    denominator = np.linalg.norm(left) * np.linalg.norm(right)
    if not denominator:
        return 0.0
    return float(np.dot(left, right) / denominator)


def _review(input_path: Path, groundtruth_path: Path) -> list[dict]:
    from app.data.embedding_client import embed_documents
    from app.data.elasticsearch_client import get_keywords

    rows = _rows(input_path, "retrieval_evaluation")
    groundtruth = {
        row["gt_id"]: row.get("expected_answer") or ""
        for row in _rows(groundtruth_path, "groundtruth")
    }
    evidence = []
    pairs = []
    for row in rows:
        sources = json.loads(row.get("sources_json") or "[]")
        evidence_text = " ".join(source.get("preview") or "" for source in sources)
        expected_answer = groundtruth.get(row["gt_id"], "")
        evidence.append((sources, evidence_text))
        pairs.append((expected_answer, evidence_text))

    vectors = np.asarray(
        embed_documents([text for pair in pairs for text in pair]),
        dtype=float,
    )
    reviewed = []
    for index, row in enumerate(rows):
        expected_answer, evidence_text = pairs[index]
        answer_vector = vectors[index * 2]
        evidence_vector = vectors[index * 2 + 1]
        similarity = _cosine(answer_vector, evidence_vector)
        answer_terms = set(get_keywords(expected_answer))
        evidence_terms = set(get_keywords(evidence_text))
        coverage = (
            len(answer_terms & evidence_terms) / len(answer_terms)
            if answer_terms
            else 0.0
        )
        sources = evidence[index][0]
        has_support_document = any(
            source.get("doc_name") in SUPPORT_DOCUMENTS
            for source in sources
        )
        top_is_support_document = row.get("top_doc_name") in SUPPORT_DOCUMENTS
        rerank_score = float(row.get("top_rerank_score") or -100)

        if (
            (
                top_is_support_document
                and rerank_score >= 0
                and (similarity >= 0.25 or coverage >= 0.45)
            )
            or similarity >= 0.58
            or (similarity >= 0.48 and coverage >= 0.60)
        ):
            verdict = "HOP_LY"
            reason = "Evidence co du tin hieu de tao cau tra loi."
        elif (
            has_support_document
            and (similarity >= 0.24 or coverage >= 0.38)
        ) or (similarity >= 0.35 and rerank_score >= 0):
            verdict = "CAN_XEM"
            reason = "Co evidence lien quan nhung top chunk hoac do day du chua tot."
        else:
            verdict = "KHONG_HOP_LY"
            reason = "Top evidence khong chua du y de tra loi cau hoi."

        if row["gt_id"] in MANUAL_OVERRIDES:
            verdict, reason = MANUAL_OVERRIDES[row["gt_id"]]

        reviewed.append({
            **row,
            "expected_answer_for_content_check": expected_answer,
            "answer_evidence_similarity": round(similarity, 4),
            "answer_keyword_coverage": round(coverage, 4),
            "reviewed_verdict": verdict,
            "reviewed_reason": reason,
            "mapping_source_used": "No",
            "gemini_calls": 0,
        })
    return reviewed


def _write(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "reviewed"
    headers = list(rows[0]) if rows else []
    sheet.append(headers)
    fills = {
        "HOP_LY": PatternFill("solid", fgColor="C6EFCE"),
        "CAN_XEM": PatternFill("solid", fgColor="FFEB9C"),
        "KHONG_HOP_LY": PatternFill("solid", fgColor="FFC7CE"),
    }
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
        cell = sheet.cell(sheet.max_row, headers.index("reviewed_verdict") + 1)
        cell.fill = fills[row["reviewed_verdict"]]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(
            max(max(len(str(cell.value or "")) for cell in column) + 2, 12),
            70,
        )

    summary = workbook.create_sheet("summary")
    summary.append(["metric", "value"])
    summary.append(["total_questions", len(rows)])
    for verdict in ("HOP_LY", "CAN_XEM", "KHONG_HOP_LY"):
        summary.append([
            verdict,
            sum(1 for row in rows if row["reviewed_verdict"] == verdict),
        ])
    summary.append(["gemini_calls", 0])
    summary.append(["mapping_source_used", "No"])
    summary.append([
        "content_reference",
        "Expected answer text only; source file/location ignored.",
    ])
    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 64
    workbook.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input",
        type=Path,
        default=ROOT_DIR / "storage" / "local_documents_eval" / "retrieval_80_questions.xlsx",
    )
    parser.add_argument(
        "--groundtruth",
        type=Path,
        default=ROOT_DIR / "groundtruth_pcntt_mapping_80_questions.xlsx",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=ROOT_DIR / "storage" / "local_documents_eval" / "retrieval_80_questions_reviewed.xlsx",
    )
    args = parser.parse_args()
    rows = _review(args.input, args.groundtruth)
    _write(args.output, rows)
    counts = {
        verdict: sum(1 for row in rows if row["reviewed_verdict"] == verdict)
        for verdict in ("HOP_LY", "CAN_XEM", "KHONG_HOP_LY")
    }
    print(json.dumps({"output": str(args.output), **counts}, ensure_ascii=False))


if __name__ == "__main__":
    main()
