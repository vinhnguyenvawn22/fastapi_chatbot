from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from create_pcntt_groundtruth_excel import extract_rows


OUTPUT_XLSX = Path("cham_diem_2_chatbot_pcntt_80_cau.xlsx")


HEADERS_GT = [
    "ID",
    "Câu hỏi test",
    "Nguồn đúng mong đợi",
    "Vị trí căn cứ",
    "Đáp án chuẩn rút gọn",
    "Ý bắt buộc phải có",
    "Keyword bắt buộc",
    "Keyword cấm",
]

HEADERS_SCORE = [
    *HEADERS_GT,
    "Bot 1 - Câu trả lời",
    "Bot 1 - Nguồn",
    "Bot 1 đúng nguồn (0-3)",
    "Bot 1 đúng trọng tâm (0-3)",
    "Bot 1 đủ ý (0-2)",
    "Bot 1 dễ hiểu (0-1)",
    "Bot 1 không bịa (0-1)",
    "Bot 1 tổng điểm",
    "Bot 1 nhận xét lỗi",
    "Bot 2 - Câu trả lời",
    "Bot 2 - Nguồn",
    "Bot 2 đúng nguồn (0-3)",
    "Bot 2 đúng trọng tâm (0-3)",
    "Bot 2 đủ ý (0-2)",
    "Bot 2 dễ hiểu (0-1)",
    "Bot 2 không bịa (0-1)",
    "Bot 2 tổng điểm",
    "Bot 2 nhận xét lỗi",
    "Bot tốt hơn",
    "Chênh lệch",
    "Ghi chú tester",
]


def clean(text: str) -> str:
    return " ".join(str(text or "").replace("\n", " ").split())


def must_have(row: dict) -> str:
    keywords = [item.strip() for item in row["keywords"].split(",") if item.strip()]
    answer = clean(row["expected_answer"])
    short_answer = answer[:220].rstrip()
    parts = []
    if short_answer:
        parts.append(short_answer)
    if keywords:
        parts.append("Từ khóa trọng tâm: " + "; ".join(keywords[:6]))
    return "; ".join(parts)


def forbidden_keywords(row: dict) -> str:
    topic = row["topic"]
    audience = row["audience"]
    forbidden = ["trả lời không có căn cứ khi đáp án chuẩn có thông tin", "nguồn không liên quan"]
    if audience == "Sinh viên":
        forbidden.extend(["cán bộ giảng viên nếu câu hỏi chỉ dành cho sinh viên", "nguồn CBGV sai đối tượng"])
    if audience == "Cán bộ giảng viên":
        forbidden.extend(["sinh viên nếu câu hỏi chỉ dành cho CBGV", "nguồn SV sai đối tượng"])
    if "Thủ tục" in topic or row["answer_type"] == "procedure":
        forbidden.append("trả lời quy chế chung nhưng không nêu thao tác hệ thống")
    return "; ".join(forbidden)


def row_to_gt(row: dict) -> list:
    return [
        row["gt_id"].replace("PCNTT-GT-", "PCNTT"),
        row["question_original"],
        row["source_file_name_expected"],
        row["required_location"],
        row["expected_answer"],
        must_have(row),
        row["keywords"],
        forbidden_keywords(row),
    ]


def style_sheet(ws, header_fill="D9EAF7"):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor=header_fill)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def auto_width(ws, widths: dict[int, int] | None = None):
    widths = widths or {}
    for col_idx in range(1, ws.max_column + 1):
        if col_idx in widths:
            ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx]
            continue
        max_len = 10
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, values_only=True):
            for value in cell:
                max_len = max(max_len, min(60, len(clean(value)) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len


def create_workbook(rows: list[dict]):
    wb = Workbook()

    ws_gt = wb.active
    ws_gt.title = "GroundTruth"
    ws_gt.append(HEADERS_GT)
    for row in rows:
        ws_gt.append(row_to_gt(row))
    style_sheet(ws_gt)
    auto_width(
        ws_gt,
        {
            1: 12,
            2: 45,
            3: 42,
            4: 24,
            5: 70,
            6: 60,
            7: 45,
            8: 45,
        },
    )

    ws_score = wb.create_sheet("ChamDiem_2_Bot")
    ws_score.append(HEADERS_SCORE)
    for index, row in enumerate(rows, start=2):
        gt_values = row_to_gt(row)
        ws_score.append(
            [
                *gt_values,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                f"=SUM(K{index}:O{index})",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                f"=SUM(T{index}:X{index})",
                "",
                f'=IF(P{index}>Y{index},"Bot 1",IF(Y{index}>P{index},"Bot 2","Hòa"))',
                f"=ABS(P{index}-Y{index})",
                "",
            ]
        )
    style_sheet(ws_score, "E2F0D9")
    auto_width(
        ws_score,
        {
            1: 12,
            2: 45,
            3: 42,
            4: 24,
            5: 60,
            6: 50,
            7: 38,
            8: 38,
            9: 60,
            10: 35,
            17: 38,
            18: 60,
            19: 35,
            26: 38,
            27: 15,
            28: 12,
            29: 30,
        },
    )

    ws_sum = wb.create_sheet("TongHop")
    total_rows = len(rows)
    total_max = total_rows * 10
    ws_sum.append(["Chỉ số", "Bot 1", "Bot 2", "Ghi chú", None])
    summary_rows = [
        ["Tổng điểm /%s" % total_max, "=SUM(ChamDiem_2_Bot!P2:P81)", "=SUM(ChamDiem_2_Bot!Y2:Y81)", f"Tổng điểm {total_rows} câu", None],
        ["Điểm trung bình /10", "=AVERAGE(ChamDiem_2_Bot!P2:P81)", "=AVERAGE(ChamDiem_2_Bot!Y2:Y81)", "Trung bình điểm", None],
        ["Số câu >= 8 điểm", '=COUNTIF(ChamDiem_2_Bot!P2:P81,">=8")', '=COUNTIF(ChamDiem_2_Bot!Y2:Y81,">=8")', "Câu trả lời tốt", None],
        ["Số câu 5-7.9 điểm", '=COUNTIFS(ChamDiem_2_Bot!P2:P81,">=5",ChamDiem_2_Bot!P2:P81,"<8")', '=COUNTIFS(ChamDiem_2_Bot!Y2:Y81,">=5",ChamDiem_2_Bot!Y2:Y81,"<8")', "Câu trung bình", None],
        ["Số câu <= 4 điểm", '=COUNTIF(ChamDiem_2_Bot!P2:P81,"<=4")', '=COUNTIF(ChamDiem_2_Bot!Y2:Y81,"<=4")', "Lỗi nặng", None],
        ["Số câu thắng", '=COUNTIF(ChamDiem_2_Bot!AA2:AA81,"Bot 1")', '=COUNTIF(ChamDiem_2_Bot!AA2:AA81,"Bot 2")', "So từng câu", None],
        ["Số câu hòa", '=COUNTIF(ChamDiem_2_Bot!AA2:AA81,"Hòa")', None, "Hai bot cùng điểm", None],
        ["Chênh lệch tổng", "=B2-C2", "=C2-B2", "Dương nghĩa là bot ở cột đó cao hơn", None],
        ["Bot điểm cao hơn", '=IF(B2>C2,"Bot 1",IF(C2>B2,"Bot 2","Hòa"))', None, "Kết luận theo tổng điểm", None],
    ]
    for row in summary_rows:
        ws_sum.append(row)
    style_sheet(ws_sum, "FCE4D6")
    auto_width(ws_sum, {1: 28, 2: 18, 3: 18, 4: 45})

    ws_guide = wb.create_sheet("HuongDan_Rubric")
    guide_rows = [
        ["Hướng dẫn chấm điểm", None],
        ["1. Dán câu trả lời và nguồn của Bot 1/Bot 2 vào sheet ChamDiem_2_Bot.", None],
        ["2. Nhập điểm từng tiêu chí: đúng nguồn 0-3, đúng trọng tâm 0-3, đủ ý 0-2, dễ hiểu 0-1, không bịa 0-1.", None],
        ["3. Cột tổng điểm tự cộng. Sheet TongHop tự tổng hợp điểm hai bot.", None],
        [None, None],
        ["Rubric", "Điểm tối đa"],
        ["Đúng nguồn tài liệu, đúng Điều/Mục", 3],
        ["Đúng trọng tâm câu hỏi", 3],
        ["Đủ ý bắt buộc", 2],
        ["Dễ hiểu, trình bày tốt", 1],
        ["Không bịa, không thêm thông tin sai", 1],
        [None, None],
        ["Gợi ý chấm đúng nguồn", "3 điểm nếu đúng file và đúng vị trí; 2 điểm nếu đúng file nhưng thiếu vị trí; 1 điểm nếu cùng nhóm nguồn nhưng chưa đúng file; 0 điểm nếu sai nguồn."],
        ["Gợi ý chấm đúng trọng tâm", "3 điểm nếu trả lời trực tiếp câu hỏi; 2 điểm nếu đúng nhưng lan man; 1 điểm nếu chỉ liên quan một phần; 0 điểm nếu lạc đề."],
        ["Gợi ý chấm đủ ý", "Dựa vào cột Ý bắt buộc phải có và Đáp án chuẩn rút gọn."],
        ["Keyword cấm", "Nếu câu trả lời chứa keyword cấm hoặc dùng sai đối tượng, cần trừ điểm đúng nguồn/không bịa tùy mức độ."],
        ["Lưu ý", "Bộ này sinh từ PCNTT_MAPPING_FILE.docx, gồm đúng 80 câu hỏi gốc, không thêm biến thể."],
    ]
    for row in guide_rows:
        ws_guide.append(row)
    style_sheet(ws_guide, "FFF2CC")
    auto_width(ws_guide, {1: 45, 2: 90})

    wb.save(OUTPUT_XLSX)


if __name__ == "__main__":
    source_rows = extract_rows()
    create_workbook(source_rows)
    print(OUTPUT_XLSX.resolve())
    print(f"rows={len(source_rows)}")
