from pathlib import Path
import re

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_DOCX = Path("documents/nghiep_vu/PCNTT_MAPPING_FILE.docx")
OUTPUT_XLSX = Path("groundtruth_pcntt_mapping_80_questions.xlsx")


def clean(text: str) -> str:
    return " ".join(str(text or "").replace("\n", " ").split())


def source_type_from_file_id(file_id: str) -> str:
    if file_id == "PCNTT_FILE_01":
        return "business_faq_mapping"
    return "business_document"


def infer_audience(file_id: str, table_index: int, question: str) -> str:
    normalized = question.lower()
    if file_id == "PCNTT_FILE_01":
        return "Cán bộ giảng viên và sinh viên"
    if file_id == "PCNTT_FILE_02" or table_index == 2:
        return "Sinh viên"
    if file_id == "PCNTT_FILE_03" or table_index == 3:
        return "Cán bộ giảng viên"
    if "giảng viên" in normalized or "cán bộ" in normalized:
        return "Cán bộ giảng viên"
    if "sinh viên" in normalized:
        return "Sinh viên"
    return "Không xác định"


def source_name_from_file_id(file_id: str, table_index: int) -> tuple[str, str]:
    if file_id == "PCNTT_FILE_01":
        return "2026.03.03.ChatbotAI_CBGV_SV_V4.docx", ""
    if file_id == "PCNTT_FILE_02" or table_index == 2:
        return (
            "2026.03.25.AI_HDSD TREN WEB SUPPORT SV.docx",
            "PCNTT_FILE_02 không xuất hiện trong bảng danh mục file nguồn; suy luận từ bảng FAQ sinh viên.",
        )
    if file_id == "PCNTT_FILE_03" or table_index == 3:
        return "2026.03.25.AI_HDSD TREN WEB SUPPORT CBGV.docx", ""
    return "", "Không có mapping file nguồn rõ ràng."


def classify_topic(question: str, keywords: str, location: str) -> str:
    text = f"{question} {keywords} {location}".lower()
    rules = [
        ("Tin tức - Thông báo", ["tin tức", "thông báo", "tiêu điểm", "bài viết"]),
        ("Kết quả học tập", ["kết quả học tập", "xem điểm", "điểm học kỳ", "điểm thành phần"]),
        ("Lịch học - Lịch thi", ["lịch học", "lịch thi", "thời khóa biểu"]),
        ("Học phí", ["học phí", "công nợ", "thanh toán"]),
        ("Thủ tục hành chính", ["thủ tục hành chính", "hồ sơ", "phê duyệt", "trình duyệt"]),
        ("Khối lượng công tác giảng viên", ["khối lượng", "công tác giảng viên", "coi thi", "chấm thi"]),
        ("Email/LMS", ["email", "lms", "mật khẩu", "xác minh"]),
        ("Thi/Kiểm tra", ["thi", "khảo thí", "phúc khảo"]),
        ("Lớp học phần", ["lớp học phần", "danh sách lớp"]),
    ]
    for topic, terms in rules:
        if any(term in text for term in terms):
            return topic
    return "Nghiệp vụ Web Support"


def answer_type(question: str) -> str:
    q = question.lower()
    if any(term in q for term in ["làm thế nào", "cách", "truy cập", "vào đâu", "xem", "kiểm tra"]):
        return "procedure"
    if any(term in q for term in ["gồm", "bao gồm", "những gì", "dữ liệu nào"]):
        return "list"
    if any(term in q for term in ["là gì", "dùng để làm gì", "vai trò"]):
        return "definition"
    if any(term in q for term in ["có thể", "không"]):
        return "yes_no"
    return "short_answer"


def paraphrase_questions(question: str, keywords: str) -> list[str]:
    q = clean(question)
    variants = [q]

    if re.match(r"(?i)^làm thế nào để ", q):
        body = re.sub(r"(?i)^làm thế nào để\s+", "", q).rstrip("?")
        body = re.sub(r"(?i)^tôi\s+", "", body)
        variants.append(f"Cách {body} như thế nào?")
        variants.append(f"Tôi cần {body} thì làm sao?")
    elif re.match(r"(?i)^làm cách nào để ", q):
        body = re.sub(r"(?i)^làm cách nào để\s+", "", q).rstrip("?")
        body = re.sub(r"(?i)^sinh viên\s+", "", body)
        variants.append(f"Cách {body} như thế nào?")
        variants.append(f"Em muốn {body} thì làm sao?")
    elif " là gì?" in q.lower():
        variants.append(re.sub(r"(?i)\s+là gì\?$", " có ý nghĩa gì?", q))
        variants.append("Cho tôi biết " + q[0].lower() + q[1:])
    elif "dùng để làm gì" in q.lower():
        variants.append(re.sub(r"(?i)dùng để làm gì\\?", "có chức năng gì?", q))
        variants.append("Cho tôi biết " + q[0].lower() + q[1:])
    elif "có thể" in q.lower() and q.endswith("?"):
        variants.append(q.replace("có thể", "được", 1))
    elif "trên Web Support" in q:
        variants.append(q.replace("trên Web Support", "trên hệ thống support", 1))

    for old, new in (("sinh viên", "em"), ("giảng viên", "thầy cô")):
        if old in q.lower() and len(variants) < 3:
            variants.append(re.sub(old, new, q, count=1, flags=re.IGNORECASE))

    keyword_list = [item.strip() for item in keywords.split(",") if item.strip()]
    if keyword_list and len(variants) < 3:
        variants.append(f"Tôi muốn hỏi về {keyword_list[0]} thì thực hiện như thế nào?")

    if len(variants) < 3:
        variants.append(f"Cho tôi hướng dẫn về nội dung: {q.rstrip('?')}.")

    deduped = []
    seen = set()
    for item in variants:
        item = clean(item)
        key = item.lower()
        if item and key not in seen:
            deduped.append(item)
            seen.add(key)
    return deduped[:3]


def extract_rows():
    document = Document(INPUT_DOCX)
    rows = []

    for table_index, table in enumerate(document.tables[1:], start=1):
        for row in table.rows[1:]:
            values = [clean(cell.text) for cell in row.cells]
            if len(values) < 6:
                continue
            stt, file_id, question, expected, location, keywords = values[:6]
            if not question or not expected:
                continue
            if not re.search(r"\d+", stt):
                continue
            source_file, note = source_name_from_file_id(file_id, table_index)
            audience = infer_audience(file_id, table_index, question)
            topic = classify_topic(question, keywords, location)
            rows.append(
                {
                    "gt_id": f"PCNTT-GT-{len(rows) + 1:03d}",
                    "table_index": table_index,
                    "stt_in_table": stt,
                    "topic": topic,
                    "audience": audience,
                    "source_type_expected": source_type_from_file_id(file_id),
                    "source_file_id": file_id,
                    "source_file_name_expected": source_file,
                    "required_location": location,
                    "keywords": keywords,
                    "question_original": question,
                    "expected_answer": expected,
                    "answer_type": answer_type(question),
                    "must_cite_source": "Có",
                    "scoring_hint": (
                        "Đạt nếu câu trả lời đúng ý chính, không trái expected_answer, "
                        "và trích đúng file/vị trí nguồn."
                    ),
                    "note": note,
                }
            )
    return rows


def write_header(ws, headers):
    ws.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def auto_width(ws, max_width=70):
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 12
        for cell in column_cells:
            value = clean(cell.value)
            if value:
                width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def create_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "groundtruth"

    groundtruth_headers = [
        "gt_id",
        "topic",
        "audience",
        "question_original",
        "expected_answer",
        "chatbot_1",
        "chatbot_2",
        "source_type_expected",
        "source_file_id",
        "source_file_name_expected",
        "required_location",
        "keywords",
        "answer_type",
        "must_cite_source",
        "scoring_hint",
        "note",
    ]
    write_header(ws, groundtruth_headers)
    for row in rows:
        ws.append([row.get(header, "") for header in groundtruth_headers])
    auto_width(ws)

    ws_map = wb.create_sheet("file_mapping")
    map_headers = ["source_file_id", "source_file_name_expected", "audience", "note"]
    write_header(ws_map, map_headers)
    mappings = {}
    for row in rows:
        key = (row["source_file_id"], row["source_file_name_expected"], row["audience"], row["note"])
        mappings[key] = True
    for key in sorted(mappings):
        ws_map.append(list(key))
    auto_width(ws_map)

    ws_readme = wb.create_sheet("README")
    ws_readme.append(["Mục", "Nội dung"])
    ws_readme.append(["Nguồn dữ liệu", str(INPUT_DOCX)])
    ws_readme.append(["Số FAQ gốc", len(rows)])
    ws_readme.append(["Số câu test", len(rows)])
    ws_readme.append([
        "Cách dùng",
        "Dùng sheet groundtruth để gửi từng câu question_original cho 2 chatbot. "
        "Điền câu trả lời vào hai cột chatbot_1 và chatbot_2, sau đó so sánh với expected_answer, "
        "source_file_name_expected và required_location.",
    ])
    ws_readme.append([
        "Cách chấm đề xuất",
        "Đúng nội dung 0-6 điểm, đúng nguồn 0-3 điểm, diễn đạt rõ ràng 0-1 điểm.",
    ])
    ws_readme.append([
        "Lưu ý",
        "PCNTT_FILE_02 không có trong bảng danh mục file nguồn của DOCX, nhưng xuất hiện trong bảng FAQ sinh viên; file nguồn được suy luận là Web Support SV.",
    ])
    auto_width(ws_readme)

    wb.save(OUTPUT_XLSX)


if __name__ == "__main__":
    rows = extract_rows()
    create_excel(rows)
    print(OUTPUT_XLSX.resolve())
    print(f"groundtruth_rows={len(rows)}")
